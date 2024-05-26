import os
import sys
import openpyxl
from schema.texts import Client_Model_1

class WriteExcel:
    file_path ='worksheets/sample_sheet.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['履歴書']
    merged_cells = sheet.merged_cells.ranges

    def __init__(self, output:Client_Model_1):
        self.output = output


    def insert_name_furigana(self):
        cell = "E5"
        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col) 
                top_left_cell.value = self.output.phonetic_character 
                break
            else:
                self.sheet[cell].value = self.output.phonetic_character
    
    def insert_name(self):
        cell = "E7"
        family_name = self.output.name.family_name
        first_name = self.output.name.first_name
        full_name = f"{family_name} {first_name}"

        cell_found = False
        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col) 
                top_left_cell.value = full_name
                cell_found = True
                break
        if not cell_found:
            self.sheet[cell].value = full_name
    
    def insert_birth_date(self):
        cell = "E10"
        if '/' in self.output.birth_date:
            text_to_split = "/"
        elif '-' in self.output.birth_date:
            text_to_split = "-"
        year = self.output.birth_date.split(text_to_split)[0]
        month = self.output.birth_date.split(text_to_split)[1]
        date = self.output.birth_date.split(text_to_split)[2]

        cell_found = False
        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = f"{year}年{month}月{date}日生 (満{self.output.age}歳)"
                cell_found = True
                break
        if not cell_found:
            self.sheet[cell].value = f"{year}年{month}月{date}日生 (満{self.output.age}歳)"
    
    def insert_gender(self):
        cell = "F10"
        if self.output.gender == "男性":
            gender = "男"
        else:
            gender = "女"

        cell_found = False
        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col) 
                top_left_cell.value = gender
                cell_found = True
                break

        if not cell_found:    
            self.sheet[cell].value = gender

    def insert_address_furigana(self):
        cell = "D13"
        cell_found = False
        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col) 
                top_left_cell.value = self.output.full_address.phonetic_character
                cell_found = True
                break
        if not cell_found:    
            self.sheet[cell].value = self.output.full_address.phonetic_character
    
    def insert_address(self):
        cell = "D15"
        cell_found = False
        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col) 
                top_left_cell.value = f"{self.output.full_address.postal_code}\n\n{self.output.full_address.address}"
                cell_found = True
                break
        if not cell_found:
            self.sheet[cell].value =  f"{self.output.full_address.postal_code}\n\n{self.output.full_address.address}"

    def insert_phone_number(self):
        cell = "G13"
        cell_found = False
        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col) 
                top_left_cell.value = f"電話 \n\n0{self.output.phone_number}"
                cell_found = True
                break
        if not cell_found:
            self.sheet[cell].value = f"電話 \n\n0{self.output.phone_number}"
    
    def insert_email(self):
        cell = "G15"
        cell_found = False
        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col) 
                top_left_cell.value = f"メール {self.output.email}"
                cell_found = True
                break
        if not cell_found:
            self.sheet[cell].value = self.output.email

            
    def insert_educational_background_year(self):
            if len(self.output.full_school_info) >= 2:
                cells = ["B26","B27","B28","B29","B30","B31","B32"]
            else:
                cells = ["B26"]

            cell_found = False
            for school_info in self.output.full_school_info:
                for cell in cells:
                    for merged_range in self.merged_cells:
                        if cell in merged_range:
                            top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                            top_left_cell.value = school_info.graduation_year
                            cell_found = True
                            break
                    if not cell_found:
                        self.sheet[cell].value = school_info.graduation_year
                        
    def insert_educational_background_month(self):
            if len(self.output.full_school_info) >= 2:
                cells = ["C26","C27","C28","C29","C30","C31","C32"]
            else:
                cells = ["C26"]
            cell_found = False
            for school_info in self.output.full_school_info:
                for cell in cells:
                    for merged_range in self.merged_cells:
                        if cell in merged_range:
                            top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                            top_left_cell.value = school_info.graduation_month
                            cell_found = True
                            break
                    if not cell_found:
                        self.sheet[cell].value = school_info.graduation_month
                            
                        
    def insert_educational_background_school_name(self):
            if len(self.output.full_school_info) >= 2:
                cells = ["E26","E27","E28","E29","E30","E31","E32"]
            else:
                cells = ["E26"]
            cell_found = False
            for school_info in self.output.full_school_info:
                for cell in cells:
                    for merged_range in self.merged_cells:
                        if cell in merged_range:
                            top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                            top_left_cell.value = school_info.school_name
                            cell_found = True
                            break
            if not cell_found:
                self.sheet[cell].value = school_info.school_name

    def insert_employment_history_year(self):
            cells = ["B36", "B38", "B40", "B42", "B44", "B46"] if len(self.output.full_employement_history) >= 2 else ["B36", "B38"]

            employment_history = list(self.output.full_employement_history)

            i = 0
            while i < len(employment_history) and i*2 < len(cells):
                employment_info = employment_history[i]
                start_cell = cells[i*2]  # 開始年用のセル

                # マージされたセルの最初のセルを見つける処理
                for merged_range in self.merged_cells:
                    if start_cell in merged_range:
                        # マージされた範囲の最初のセルを取得
                        start_row = merged_range.min_row
                        start_col = merged_range.min_col
                        break
                else:
                    # マージされていない場合はそのまま使用
                    start_row, start_col = self.sheet[start_cell].row, self.sheet[start_cell].column

                # 開始年を設定
                self.sheet.cell(row=start_row, column=start_col).value = employment_info.year_started

                if (i*2 + 1) < len(cells):  # 次のセルがあれば
                    end_cell = cells[i*2 + 1]
                    # マージされたセルの最初のセルを見つける
                    for merged_range in self.merged_cells:
                        if end_cell in merged_range:
                            # マージされた範囲の最初のセルを取得
                            end_row = merged_range.min_row
                            end_col = merged_range.min_col
                            break
                    else:
                        # マージされていない場合はそのまま使用
                        end_row, end_col = self.sheet[end_cell].row, self.sheet[end_cell].column

                    # 終了年を設定
                    self.sheet.cell(row=end_row, column=end_col).value = employment_info.year_ended

                i += 1  # 次の履歴へ

    def insert_employment_history_month(self):
            cells = ["C36", "C38", "C40", "C42", "C44", "C46"] if len(self.output.full_employement_history) >= 2 else ["C36", "C38"]

            employment_history = list(self.output.full_employement_history)

            i = 0
            while i < len(employment_history) and i*2 < len(cells):
                employment_info = employment_history[i]
                start_cell = cells[i*2]  # 開始年用のセル

                # マージされたセルの最初のセルを見つける処理
                for merged_range in self.merged_cells:
                    if start_cell in merged_range:
                        # マージされた範囲の最初のセルを取得
                        start_row = merged_range.min_row
                        start_col = merged_range.min_col
                        break
                else:
                    # マージされていない場合はそのまま使用
                    start_row, start_col = self.sheet[start_cell].row, self.sheet[start_cell].column

                # 開始年を設定
                self.sheet.cell(row=start_row, column=start_col).value = employment_info.month_started

                if (i*2 + 1) < len(cells):  # 次のセルがあれば
                    end_cell = cells[i*2 + 1]
                    # マージされたセルの最初のセルを見つける
                    for merged_range in self.merged_cells:
                        if end_cell in merged_range:
                            # マージされた範囲の最初のセルを取得
                            end_row = merged_range.min_row
                            end_col = merged_range.min_col
                            break
                    else:
                        # マージされていない場合はそのまま使用
                        end_row, end_col = self.sheet[end_cell].row, self.sheet[end_cell].column

                    # 終了年を設定
                    self.sheet.cell(row=end_row, column=end_col).value = employment_info.month_ended

                i += 1  # 次の履歴へ

    def insert_employment_history_name(self):
            cells = ["E36", "E38", "E40", "E42", "E44", "E46"] if len(self.output.full_employement_history) >= 2 else ["E36", "E38"]

            employment_history = list(self.output.full_employement_history)

            i = 0
            while i < len(employment_history) and i*2 < len(cells):
                employment_info = employment_history[i]
                start_cell = cells[i*2]  # 開始年用のセル

                # マージされたセルの最初のセルを見つける処理
                for merged_range in self.merged_cells:
                    if start_cell in merged_range:
                        # マージされた範囲の最初のセルを取得
                        start_row = merged_range.min_row
                        start_col = merged_range.min_col
                        break
                else:
                    # マージされていない場合はそのまま使用
                    start_row, start_col = self.sheet[start_cell].row, self.sheet[start_cell].column

                # 開始年を設定
                self.sheet.cell(row=start_row, column=start_col).value = employment_info.company_name

                if (i*2 + 1) < len(cells):  # 次のセルがあれば
                    end_cell = cells[i*2 + 1]
                    # マージされたセルの最初のセルを見つける
                    for merged_range in self.merged_cells:
                        if end_cell in merged_range:
                            # マージされた範囲の最初のセルを取得
                            end_row = merged_range.min_row
                            end_col = merged_range.min_col
                            break
                    else:
                        # マージされていない場合はそのまま使用
                        end_row, end_col = self.sheet[end_cell].row, self.sheet[end_cell].column

                    # 終了年を設定
                    self.sheet.cell(row=end_row, column=end_col).value = employment_info.company_name

                i += 1  # 次の履歴へ
    
    def insert_qualification_year(self):
        cells = ["K17", "K19", "K20", "K22", "K24", "K26"]
        qualifications = list(self.output.full_qualifications)

        for cell, qualification in zip(cells, qualifications):
            for merged_range in self.merged_cells:
                if cell in merged_range:
                    top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    top_left_cell.value = qualification.year_obtained
                    break
                else:
                    self.sheet[cell].value = qualification.year_obtained
                    break

    def insert_qualification_month(self):
        cells = ["L17", "L19", "L20", "L22", "L24", "L26"]
        qualifications = list(self.output.full_qualifications)

        for cell, qualification in zip(cells, qualifications):
            for merged_range in self.merged_cells:
                if cell in merged_range:
                    top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    top_left_cell.value = qualification.month_obtained
                    break
                else:
                    self.sheet[cell].value = qualification.month_obtained
                    break
    def insert_qualification_name(self):
        cells = ["M17", "M19", "M20", "M22", "M24", "M26"]
        qualifications = list(self.output.full_qualifications)

        for cell, qualification in zip(cells, qualifications):
            for merged_range in self.merged_cells:
                if cell in merged_range:
                    top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    top_left_cell.value = qualification.name
                    break
                else:
                    self.sheet[cell].value = qualification.name
                    break
    
    def insert_family_member(self):
        cell = "N33"

        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col) 
                top_left_cell.value = f"{self.output.family_member}人"
                break
            else:
                self.sheet[cell].value = f"{self.output.family_member}人"
    
    def insert_is_married(self):
        cell = "N37"
        if self.output.Married:
            is_married = "有"
        else:
            is_married = "無"

        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col) 
                top_left_cell.value = is_married
                break
            else:
                self.sheet[cell].value = is_married

    
    def insert_is_support_obligated(self):
        cell = "O37"
        if self.output.support_obligation:
            is_obligated = "有"
        else:
            is_obligated = "無"

        for merged_range in self.merged_cells:
            if cell in merged_range:
                top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col) 
                top_left_cell.value = is_obligated
                break
            else:
                self.sheet[cell].value = is_obligated


    def save_data(self):
        family_name = self.output.name.family_name
        first_name = self.output.name.first_name
        full_name = f"{family_name} {first_name}"
        file_path_to_save = f"worksheets/{full_name}様.xlsx"
        self.wb.save(file_path_to_save)

    def export_to_xlsx(self):
        self.insert_name_furigana()
        self.insert_name()
        self.insert_birth_date()
        self.insert_gender()
        self.insert_address_furigana()
        self.insert_gender()
        self.insert_address()
        self.insert_phone_number()
        self.insert_email()
        self.insert_educational_background_month()
        self.insert_educational_background_year()
        self.insert_educational_background_school_name()
        self.insert_employment_history_year()
        self.insert_employment_history_month()
        self.insert_employment_history_name()
        self.insert_qualification_year()
        self.insert_qualification_month()
        self.insert_qualification_name()
        self.insert_family_member()
        self.insert_is_married()
        self.insert_is_support_obligated()
        self.save_data()
